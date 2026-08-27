from __future__ import annotations

import asyncio
import json
import logging
import math
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

from confluent_kafka import Consumer
from openpyxl import load_workbook
from superagentx.handler.base import BaseHandler
from superagentx.handler.decorators import tool
from superagentx.llm import LLMClient

from superagentx.llm.litellm import LiteLLMClient
from superagentx.llm.models import ChatCompletionParams

from superagentx_policy_engine.models.request import AuthorizationRequest
from superagentx_policy_engine.policy_engine import PolicyEngine
from superagentx_policy_engine.store.file_store import FilePolicyStore


# ============================================================
# CONFIG
# ============================================================

BOOTSTRAP_SERVERS = os.environ.get("KAFKA_BOOTSTRAP_SERVERS", "localhost:9092")
TOPIC_NAME = os.environ.get("KAFKA_TOPIC_NAME", "policy-input")
CONSUMER_GROUP_ID = os.environ.get("KAFKA_CONSUMER_GROUP_ID", "policy-evaluation-consumer")
DEFAULT_OUTPUT_DIR = os.environ.get("DEFAULT_OUTPUT_DIR", "/app/output")

ROW_NUMBER_FIELD = "__source_row_number"
RESULT_HEADERS = ("Decision", "Threat Score", "Threat Severity")
POLL_TIMEOUT_SECONDS = 1.0

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)


# ============================================================
# EXCEL
# ============================================================

class WorkbookService:
    """Create a separate processed workbook without modifying the source file."""

    @staticmethod
    def _upsert_headers(worksheet) -> dict[str, int]:
        """
        Find the last column that holds real source data and place the
        three generated result headers immediately after it, with zero
        gap. Anything sitting to the right of the source data (old result
        columns from a previous run, stray formatting, orphaned columns,
        etc.) is wiped first so the new headers can never end up with
        empty columns between them and the source data.
        """
        header_map: dict[str, int] = {}

        max_col = worksheet.max_column or 0

        # Real source data is a CONTIGUOUS run of headers starting at
        # column A. The first empty (or result-header) cell marks the end
        # of that run. Anything further right — even a non-empty cell,
        # such as stray leftover text from a previous/corrupted run — is
        # noise, not source data, and must NOT extend the boundary. This
        # avoids the case where an isolated stray value sitting far out
        # (e.g. a single leftover cell) gets mistaken for real data and
        # leaves a gap between the actual source columns and the results.
        last_source_column = 0
        for col in range(1, max_col + 1):
            value = worksheet.cell(row=1, column=col).value

            if value is not None:
                value = str(value).strip()

            if not value or value in RESULT_HEADERS:
                # End of the contiguous source-header run.
                break

            last_source_column = col

        if last_source_column == 0:
            raise ValueError(
                "No source headers found in Excel file."
            )

        result_columns = {
            "Decision": last_source_column + 1,
            "Threat Score": last_source_column + 2,
            "Threat Severity": last_source_column + 3,
        }

        # Wipe everything to the right of the source data (all rows).
        # This removes any previously generated result columns no matter
        # where they ended up, plus any stray blank/formatted columns,
        # guaranteeing the new result columns sit immediately after the
        # source data with no gap.
        if max_col > last_source_column:
            for row_number in range(1, worksheet.max_row + 1):
                for col in range(last_source_column + 1, max_col + 1):
                    worksheet.cell(row=row_number, column=col).value = None

        # Create result headers directly after the last source column.
        for header, column in result_columns.items():
            worksheet.cell(
                row=1,
                column=column,
                value=header,
            )
            header_map[header] = column

        return header_map

    def create_output_workbook(
        self,
        source_file_path: str,
        output_file_path: str,
        results: list[dict],
    ) -> str:
        source_path = Path(source_file_path)
        output_path = Path(output_file_path)

        if not source_path.exists():
            raise FileNotFoundError(f"Source Excel file not found: {source_path}")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        logger.info("Creating processed workbook source=%s output=%s results=%s", source_path, output_path, len(results))

        workbook = load_workbook(source_path)
        worksheet = workbook.active
        try:
            header_map = self._upsert_headers(worksheet)
            for result in results:
                row_number = result.get("row_number")
                if row_number is None:
                    logger.warning("Skipping result without source row number: %r", result)
                    continue
                worksheet.cell(row=int(row_number), column=header_map["Decision"], value=result.get("decision"))
                worksheet.cell(row=int(row_number), column=header_map["Threat Score"], value=result.get("threat_score"))
                worksheet.cell(row=int(row_number), column=header_map["Threat Severity"], value=result.get("threat_severity"))
            workbook.save(output_path)
        finally:
            workbook.close()

        logger.info("Processed Excel created: %s", output_path)
        return str(output_path)


# ============================================================
# POLICY EVALUATOR
# ============================================================

class PolicyEvaluator:

    # --------------------------------------------------------
    # Excel field <-> Policy fact aliases
    # --------------------------------------------------------

    FIELD_ALIASES = {
        "amount": ["amount", "transaction_amount", "payment_amount", "txn_amount", "value"],
        "transaction_id": [
            "transaction_id", "transaction_reference", "transaction_ref",
            "txn_id", "payment_id", "reference", "id",
        ],
        "country": [
            "country", "destination_country", "beneficiary_country",
            "recipient_country", "receiver_country",
        ],
        "channel": [
            "channel", "payment_channel", "transaction_channel",
            "transaction_mode", "payment_mode",
        ],
        "risk_score": ["risk_score", "score", "transaction_risk_score", "combined_risk_score"],
        "threat_score": ["threat_score", "risk_score", "score"],
        "threat_severity": ["threat_severity", "risk_level", "severity", "threat_level"],
        "beneficiary": [
            "beneficiary", "beneficiary_name", "recipient", "receiver",
            "counterparty", "counterparty_name",
        ],
        "account_id": ["account_id", "account_number", "customer_account"],
        "payment_type": ["payment_type", "transaction_type", "type"],
        "payment_mode": ["payment_mode", "transaction_mode", "channel"],
    }

    # --------------------------------------------------------
    # AuthorizationRequest.principal.type only accepts these
    # exact values. Anything else the LLM invents must be
    # mapped onto one of these or validation raises.
    # --------------------------------------------------------

    PRINCIPAL_TYPES = (
        "User", "Agent", "System", "Service", "Workflow",
        "Engine", "Handler", "Tool", "ApiClient", "Tenant",
        "Role", "Group", "Application",
    )

    PRINCIPAL_TYPE_ALIASES = {
        "user": "User",
        "customer": "User",
        "person": "User",
        "individual": "User",
        "human": "User",
        "employee": "User",
        "agent": "Agent",
        "ai_agent": "Agent",
        "system": "System",
        "service": "Service",
        "microservice": "Service",
        "workflow": "Workflow",
        "engine": "Engine",
        "handler": "Handler",
        "tool": "Tool",
        "apiclient": "ApiClient",
        "api_client": "ApiClient",
        "client": "ApiClient",
        "tenant": "Tenant",
        "role": "Role",
        "group": "Group",
        "application": "Application",
        "app": "Application",
    }

    def __init__(self, llm: LLMClient, policy_path: str):
        self.policy_file = policy_path

        # Load policy ONCE
        policy_path = self._resolve_policy_path()
        logger.info("Loading policy once from %s", policy_path)
        self.full_policy = FilePolicyStore.load(str(policy_path))
        self.full_policy_json = self.full_policy.model_dump(mode="json")
        self.all_statements = self.full_policy_json.get("statements", [])
        logger.info("Loaded %s policy statements", len(self.all_statements))

        # Build policy index ONCE
        self.policy_index = self._build_policy_index(self.all_statements)

        # LLM created ONCE
        self.llm = llm

    # ========================================================
    # PATH
    # ========================================================

    def _resolve_policy_path(self) -> Path:
        policy_path = Path(self.policy_file)
        if not policy_path.is_absolute():
            policy_path = Path(__file__).resolve().parent / policy_path
        if not policy_path.exists():
            raise FileNotFoundError(f"Policy file not found: {policy_path}")
        return policy_path

    # ========================================================
    # NORMALIZE FIELD NAME
    # ========================================================

    @staticmethod
    def _normalize_name(value: Any) -> str:
        value = str(value).strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        return value.strip("_")

    # ========================================================
    # TYPE
    # ========================================================

    @staticmethod
    def _value_type(value: Any) -> str:
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, list):
            return "array"
        if isinstance(value, dict):
            return "object"
        return "string"

    # ========================================================
    # PRINCIPAL TYPE NORMALIZATION (the crash fix)
    # ========================================================

    @classmethod
    def _normalize_principal_type(cls, value: Any) -> str:
        """Force principal.type onto one of the enum values the
        AuthorizationRequest schema accepts. Defaults to "User" for
        anything unrecognized (e.g. "Customer", "Client", empty, None)."""
        if value:
            text = str(value).strip()
            if text in cls.PRINCIPAL_TYPES:
                return text
            key = cls._normalize_name(text)
            if key in cls.PRINCIPAL_TYPE_ALIASES:
                return cls.PRINCIPAL_TYPE_ALIASES[key]
        return "User"

    # ========================================================
    # POLICY INDEX
    # ========================================================

    def _build_policy_index(self, statements: list[dict]) -> dict:
        index = {"facts": {}, "actions": {}, "resources": {}}

        for statement in statements:
            actions = statement.get("action", [])
            if isinstance(actions, str):
                actions = [actions]
            for action in actions:
                key = self._normalize_name(action)
                index["actions"].setdefault(key, []).append(statement)

            resources = statement.get("resource", [])
            if isinstance(resources, str):
                resources = [resources]
            for resource in resources:
                resource_type = resource.split(":", 1)[0] if ":" in resource else resource
                key = self._normalize_name(resource_type)
                index["resources"].setdefault(key, []).append(statement)

            condition = statement.get("condition") or {}
            for condition_type in ("all", "any"):
                rules = condition.get(condition_type, [])
                if not isinstance(rules, list):
                    continue
                for rule in rules:
                    if not isinstance(rule, dict):
                        continue
                    fact = str(rule.get("fact", ""))
                    if not fact.startswith("context.custom."):
                        continue
                    fact_name = fact.removeprefix("context.custom.")
                    fact_key = self._normalize_name(fact_name)
                    index["facts"].setdefault(fact_key, []).append(statement)

        logger.info(
            "Policy index built facts=%s actions=%s resources=%s",
            len(index["facts"]), len(index["actions"]), len(index["resources"]),
        )
        return index

    # ========================================================
    # RECORD LOOKUP
    # ========================================================

    def _record_lookup(self, record: dict) -> dict[str, Any]:
        return {self._normalize_name(key): value for key, value in record.items()}

    # ========================================================
    # GET RECORD VALUE
    # ========================================================

    def _get_record_value(self, record_lookup: dict, fact_name: str):
        fact = self._normalize_name(fact_name)

        if fact in record_lookup:
            return True, record_lookup[fact]

        for record_key, value in record_lookup.items():
            if len(fact) >= 4 and (fact in record_key or record_key in fact):
                return True, value

        for canonical, aliases in self.FIELD_ALIASES.items():
            canonical = self._normalize_name(canonical)
            normalized_aliases = {self._normalize_name(alias) for alias in aliases}
            if fact == canonical or fact in normalized_aliases:
                for possible_name in {canonical, *normalized_aliases}:
                    if possible_name in record_lookup:
                        return True, record_lookup[possible_name]

        return False, None

    # ========================================================
    # GET FACTS IN POLICY
    # ========================================================

    def _statement_fact_rules(self, statement: dict) -> list[dict]:
        result = []
        condition = statement.get("condition") or {}

        for condition_type in ("all", "any"):
            rules = condition.get(condition_type, [])
            if not isinstance(rules, list):
                continue
            for rule in rules:
                if not isinstance(rule, dict):
                    continue
                fact = str(rule.get("fact", ""))
                if not fact.startswith("context.custom."):
                    continue
                result.append({
                    "condition_type": condition_type,
                    "fact": fact.removeprefix("context.custom."),
                    "operator": rule.get("operator"),
                    "value": rule.get("value"),
                })
        return result

    # ========================================================
    # FAST CONDITION CHECK
    # ========================================================

    @staticmethod
    def _as_number(value: Any):
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            if isinstance(value, float) and math.isnan(value):
                return None
            return float(value)
        if value is None:
            return None

        text = str(value).replace(",", "")
        text = re.sub(r"[^\d.\-]", "", text)
        if not text:
            return None

        try:
            return float(text)
        except ValueError:
            return None

    def _condition_might_match(self, actual_value: Any, operator: str | None, expected_value: Any) -> bool:
        if operator is None:
            return True

        operator = str(operator).strip().lower()

        try:
            if operator in {"equal", "equals", "eq", "=="}:
                return str(actual_value).lower() == str(expected_value).lower()

            if operator in {"not_equal", "not_equals", "neq", "!="}:
                return str(actual_value).lower() != str(expected_value).lower()

            actual_number = self._as_number(actual_value)
            expected_number = self._as_number(expected_value)

            if operator in {"greater_than", "gt", ">"}:
                if actual_number is None or expected_number is None:
                    return True
                return actual_number > expected_number

            if operator in {"greater_than_or_equal", "gte", ">="}:
                if actual_number is None or expected_number is None:
                    return True
                return actual_number >= expected_number

            if operator in {"less_than", "lt", "<"}:
                if actual_number is None or expected_number is None:
                    return True
                return actual_number < expected_number

            if operator in {"less_than_or_equal", "lte", "<="}:
                if actual_number is None or expected_number is None:
                    return True
                return actual_number <= expected_number

            if operator == "in":
                if isinstance(expected_value, list):
                    normalized = {str(x).lower() for x in expected_value}
                    return str(actual_value).lower() in normalized
                return str(actual_value).lower() in str(expected_value).lower()

            if operator == "not_in":
                if isinstance(expected_value, list):
                    normalized = {str(x).lower() for x in expected_value}
                    return str(actual_value).lower() not in normalized
                return True

            if operator == "contains":
                return str(expected_value).lower() in str(actual_value).lower()

            if operator == "starts_with":
                return str(actual_value).lower().startswith(str(expected_value).lower())

            return True

        except Exception:
            return True

    # ========================================================
    # STATEMENT RELEVANCE
    # ========================================================

    def _statement_is_relevant(self, statement: dict, record_lookup: dict) -> bool:
        rules = self._statement_fact_rules(statement)
        if not rules:
            return True

        all_rules = [r for r in rules if r["condition_type"] == "all"]
        any_rules = [r for r in rules if r["condition_type"] == "any"]

        known_all_results = []
        for rule in all_rules:
            found, value = self._get_record_value(record_lookup, rule["fact"])
            if not found:
                continue
            known_all_results.append(self._condition_might_match(value, rule["operator"], rule["value"]))

        if known_all_results and not all(known_all_results):
            return False

        known_any_results = []
        for rule in any_rules:
            found, value = self._get_record_value(record_lookup, rule["fact"])
            if not found:
                continue
            known_any_results.append(self._condition_might_match(value, rule["operator"], rule["value"]))

        if known_any_results and not any(known_any_results):
            return False

        matched_fact_count = sum(
            1 for rule in rules if self._get_record_value(record_lookup, rule["fact"])[0]
        )
        return matched_fact_count > 0

    # ========================================================
    # FILTER POLICIES
    # ========================================================

    def filter_relevant_statements(self, record: dict) -> list[dict]:
        record_lookup = self._record_lookup(record)
        logger.info("Record fields: %s", list(record_lookup.keys()))

        candidates = {}

        for policy_fact, statements in self.policy_index["facts"].items():
            found, _ = self._get_record_value(record_lookup, policy_fact)
            if not found:
                continue
            for statement in statements:
                sid = str(statement.get("sid") or id(statement))
                candidates[sid] = statement

        for statement in self.all_statements:
            if not self._statement_fact_rules(statement):
                sid = str(statement.get("sid") or id(statement))
                candidates[sid] = statement

        if not candidates:
            logger.warning("No policy fact matched record columns. Using full policy set as fallback.")
            candidate_list = list(self.all_statements)
        else:
            candidate_list = list(candidates.values())

        relevant = [
            statement for statement in candidate_list
            if self._statement_is_relevant(statement, record_lookup)
        ]

        if not relevant:
            logger.warning("Relevant policy filter returned 0. Using candidate policies as fallback.")
            relevant = candidate_list

        logger.info(
            "Policy filter: total=%s candidate=%s relevant=%s",
            len(self.all_statements), len(candidate_list), len(relevant),
        )
        return relevant

    # ========================================================
    # POLICY SCHEMA
    # ========================================================

    def build_policy_schema(self, statements: list[dict]) -> dict:
        actions = set()
        resource_types = set()
        custom_facts = {}
        policy_sids = []

        for statement in statements:
            sid = statement.get("sid")
            if sid:
                policy_sids.append(sid)

            statement_actions = statement.get("action", [])
            if isinstance(statement_actions, str):
                statement_actions = [statement_actions]
            actions.update(statement_actions)

            statement_resources = statement.get("resource", [])
            if isinstance(statement_resources, str):
                statement_resources = [statement_resources]
            for resource in statement_resources:
                if ":" not in resource:
                    continue
                resource_type = resource.split(":", 1)[0]
                if resource_type != "*":
                    resource_types.add(resource_type)

            for rule in self._statement_fact_rules(statement):
                fact_name = rule["fact"]
                expected_value = rule["value"]
                custom_facts.setdefault(fact_name, {"type": self._value_type(expected_value), "conditions": []})
                custom_facts[fact_name]["conditions"].append({
                    "sid": sid,
                    "operator": rule["operator"],
                    "value": expected_value,
                })

        return {
            "actions": sorted(actions),
            "resource_types": sorted(resource_types),
            "custom_facts": custom_facts,
            "policy_sids": policy_sids,
        }

    # ========================================================
    # COMPACT PARC PROMPT
    # ========================================================

    @staticmethod
    def build_prompt(record: dict, policy_schema: dict) -> str:
        return f"""
Create one SuperAgentX AuthorizationRequest JSON object.

Return ONLY valid JSON.
No markdown.
No explanation.

Allowed action names:
{json.dumps(policy_schema["actions"])}

Allowed resource types:
{json.dumps(policy_schema["resource_types"])}

Required context.custom facts:
{json.dumps(policy_schema["custom_facts"], default=str)}

Rules:

- Return a JSON object.
- principal must be an object with id and type.
- principal.type must be exactly one of: "User", "Agent", "System",
  "Service", "Workflow", "Engine", "Handler", "Tool", "ApiClient",
  "Tenant", "Role", "Group", "Application". Use "User" for a customer
  or individual initiating the transaction.
- action must be an object with name.
- resource must be an object with type and id.
- context must be an object.
- context.custom must be an object.

- action.name must use one exact allowed action.
- resource.type must use one exact allowed resource type.

- Use transaction id, payment id, reference,
  account id, or record id as resource.id.

- Populate custom facts from the transaction.

- If a boolean value cannot be inferred, use false.
- If a numeric value cannot be inferred, use 0.
- If an array cannot be inferred, use [].
- If a string cannot be inferred, use "".

- context.risk_score must be numeric and never null.

- context.custom.threat_score must be numeric
  and never null.

- context.custom.threat_severity must be exactly:
  "low", "medium", or "high".

- context.custom.threat_reasons must be an array.

Transaction:
{json.dumps(record, default=str)}
""".strip()

    # ========================================================
    # JSON RESPONSE CLEANER
    # ========================================================

    @staticmethod
    def _clean_llm_json(content: Any) -> dict:
        if content is None:
            raise ValueError("Gemini returned empty response")

        if isinstance(content, dict):
            return content

        content = str(content).strip()
        if not content:
            raise ValueError("Gemini returned empty response")

        content = re.sub(r"^\s*```(?:json)?\s*", "", content, flags=re.IGNORECASE)
        content = re.sub(r"\s*```\s*$", "", content)

        try:
            parsed = json.loads(content)
            if not isinstance(parsed, dict):
                raise ValueError("Gemini response JSON is not an object")
            return parsed
        except json.JSONDecodeError:
            pass

        start = content.find("{")
        end = content.rfind("}")
        if start == -1 or end == -1 or end <= start:
            logger.error("Invalid Gemini response: %r", content[:1000])
            raise ValueError("Gemini did not return JSON")

        json_text = content[start:end + 1]
        try:
            parsed = json.loads(json_text)
        except json.JSONDecodeError as exc:
            logger.error("Invalid Gemini JSON: %r", json_text[:1000])
            raise ValueError("Gemini returned invalid JSON") from exc

        if not isinstance(parsed, dict):
            raise ValueError("Gemini JSON must be an object")
        return parsed

    # ========================================================
    # STRING -> DICT SAFETY
    # ========================================================

    @staticmethod
    def _ensure_dict(value: Any) -> dict:
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            value = value.strip()
            if value.startswith("{"):
                try:
                    parsed = json.loads(value)
                    if isinstance(parsed, dict):
                        return parsed
                except Exception:
                    pass
        return {}

    # ========================================================
    # NORMALIZE PARC
    # ========================================================

    def normalize_request(self, llm_context: dict, record: dict, policy_schema: dict) -> dict:
        if not isinstance(llm_context, dict):
            raise ValueError("Gemini PARC must be a JSON object")

        # ---------------------------------------------------- PRINCIPAL
        principal = self._ensure_dict(llm_context.get("principal"))
        llm_context["principal"] = principal

        if not principal.get("id"):
            principal["id"] = str(
                record.get("customer_id")
                or record.get("Customer ID")
                or record.get("account_id")
                or record.get("Account ID")
                or record.get("Account Number")
                or record.get("account_number")
                or "unknown"
            )

        # principal.type must match the AuthorizationRequest enum exactly,
        # otherwise pydantic validation raises (this was the crash).
        principal["type"] = self._normalize_principal_type(principal.get("type"))

        # ---------------------------------------------------- ACTION
        raw_action = llm_context.get("action")
        action = {"name": raw_action} if isinstance(raw_action, str) else self._ensure_dict(raw_action)
        llm_context["action"] = action

        actions = policy_schema.get("actions", [])
        if len(actions) == 1:
            action["name"] = actions[0]
        elif actions and action.get("name") not in actions:
            action["name"] = actions[0]

        # ---------------------------------------------------- RESOURCE
        raw_resource = llm_context.get("resource")
        resource = {"id": raw_resource} if isinstance(raw_resource, str) else self._ensure_dict(raw_resource)
        llm_context["resource"] = resource

        resource_types = policy_schema.get("resource_types", [])
        if len(resource_types) == 1:
            resource["type"] = resource_types[0]
        elif resource_types and resource.get("type") not in resource_types:
            resource["type"] = resource_types[0]

        if not resource.get("id"):
            resource["id"] = str(
                record.get("Transaction ID")
                or record.get("transaction_id")
                or record.get("Transaction Id")
                or record.get("Payment ID")
                or record.get("payment_id")
                or record.get("Reference")
                or record.get("reference")
                or record.get("id")
                or "unknown"
            )

        # ---------------------------------------------------- CONTEXT
        context = self._ensure_dict(llm_context.get("context"))
        llm_context["context"] = context

        custom = self._ensure_dict(context.get("custom"))
        context["custom"] = custom

        context["prompt"] = json.dumps(record, default=str)

        # ---------------------------------------------------- FILL FACTS
        record_lookup = self._record_lookup(record)
        required_facts = policy_schema.get("custom_facts", {})

        for fact_name, definition in required_facts.items():
            if fact_name in custom and custom[fact_name] is not None:
                continue

            found, actual_value = self._get_record_value(record_lookup, fact_name)
            if found:
                custom[fact_name] = actual_value
                continue

            value_type = definition.get("type", "string")
            if value_type == "boolean":
                custom[fact_name] = False
            elif value_type == "number":
                custom[fact_name] = 0
            elif value_type == "array":
                custom[fact_name] = []
            elif value_type == "object":
                custom[fact_name] = {}
            else:
                custom[fact_name] = ""

        # ---------------------------------------------------- RISK SCORE
        risk_score = (
            context.get("risk_score")
            or custom.get("risk_score")
            or custom.get("combined_risk_score")
            or custom.get("threat_score")
            or 0
        )
        parsed_risk_score = self._as_number(risk_score)
        if parsed_risk_score is None:
            parsed_risk_score = 0
        context["risk_score"] = parsed_risk_score

        # ---------------------------------------------------- THREAT SCORE
        parsed_threat = self._as_number(custom.get("threat_score"))
        if parsed_threat is None:
            parsed_threat = parsed_risk_score
        custom["threat_score"] = parsed_threat

        # ---------------------------------------------------- THREAT SEVERITY
        severity = str(custom.get("threat_severity") or "").strip().lower()
        if severity not in {"low", "medium", "high"}:
            if parsed_threat >= 70:
                severity = "high"
            elif parsed_threat >= 40:
                severity = "medium"
            else:
                severity = "low"
        custom["threat_severity"] = severity

        # ---------------------------------------------------- REASONS
        reasons = custom.get("threat_reasons")
        if isinstance(reasons, str):
            reasons = [reasons]
        elif not isinstance(reasons, list):
            reasons = []
        custom["threat_reasons"] = reasons

        logger.info(
            "Normalized PARC action=%s resource_type=%s resource_id=%s "
            "principal_type=%s risk_score=%s threat_score=%s severity=%s",
            action.get("name"), resource.get("type"), resource.get("id"),
            principal.get("type"), context.get("risk_score"),
            custom.get("threat_score"), custom.get("threat_severity"),
        )

        return llm_context

    # ========================================================
    # CREATE FILTERED POLICY
    # ========================================================

    def create_filtered_policy(self, statements: list[dict]):
        filtered_json = dict(self.full_policy_json)
        filtered_json["statements"] = statements

        policy_class = self.full_policy.__class__
        if hasattr(policy_class, "model_validate"):
            return policy_class.model_validate(filtered_json)
        return policy_class(**filtered_json)

    # ========================================================
    # EVALUATE ONE RECORD
    # ========================================================

    async def evaluate(self, record: dict) -> dict:
        relevant_statements = self.filter_relevant_statements(record)
        policy_schema = self.build_policy_schema(relevant_statements)

        logger.info("Relevant policy SIDs: %s", policy_schema["policy_sids"])
        logger.info("Generating PARC using %s relevant policies", len(relevant_statements))

        prompt = self.build_prompt(record, policy_schema)

        llm_response = await self.llm.achat_completion(
            chat_completion_params=ChatCompletionParams(
                temperature=0,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Return only one valid JSON object for a SuperAgentX "
                            "AuthorizationRequest. context and context.custom "
                            "must be JSON objects. principal.type must be one of "
                            "the allowed enum values given in the prompt."
                        ),
                    },
                    {"role": "user", "content": prompt},
                ],
            )
        )

        content = llm_response.choices[0].message.content
        llm_context = self._clean_llm_json(content)

        logger.info(
            "LLM PARC types: principal=%s action=%s resource=%s context=%s",
            type(llm_context.get("principal")).__name__,
            type(llm_context.get("action")).__name__,
            type(llm_context.get("resource")).__name__,
            type(llm_context.get("context")).__name__,
        )

        llm_context = self.normalize_request(llm_context, record, policy_schema)

        try:
            request = AuthorizationRequest(**llm_context)
        except Exception:
            logger.error(
                "AuthorizationRequest validation failed.\nNormalized PARC:\n%s",
                json.dumps(llm_context, indent=2, default=str),
            )
            raise

        filtered_policy = self.create_filtered_policy(relevant_statements)
        engine = PolicyEngine()
        await engine.add_policy_document(filtered_policy)

        response = await engine.authorize(request)
        custom = request.context.custom

        return {
            "decision": response.decision,
            "threat_score": custom.get("threat_score"),
            "threat_severity": custom.get("threat_severity"),
            "threat_reasons": custom.get("threat_reasons", []),
            "policy_sid": response.policy_sid,
            "matched_policy_sids": policy_schema["policy_sids"],
        }



# ============================================================
# KAFKA CONSUMER HANDLER
# ============================================================

class KafkaConsumerHandler(BaseHandler):
    """
    SuperAgentX Kafka consumer handler for banking policy evaluation.

    The handler consumes transaction batches from Kafka, evaluates each
    transaction using PolicyEvaluator, collects Decision / Threat Score / Threat Severity results, commits each Kafka
    message after successful batch processing, and creates a separate output
    Excel file only after all expected batches and records are complete.

    The consumer does not import any constants or code from the producer.
    Row information is read directly from the incoming Kafka record using
    ``ROW_NUMBER_FIELD`` defined in this consumer module.
    """

    def __init__(
        self,
        llm: LLMClient,
        policy_path: str,
        bootstrap_servers: Optional[str] = None,
        topic_name: Optional[str] = None,
        consumer_group_id: Optional[str] = None,
        output_dir_path: Optional[str] = None,
        poll_timeout_seconds: float = POLL_TIMEOUT_SECONDS,
        **kwargs,
    ):
        """
        Initialize the Kafka policy consumer handler.

        Args:
            bootstrap_servers: Kafka bootstrap server address.
            topic_name: Kafka topic containing policy evaluation batches.
            consumer_group_id: Kafka consumer group ID.
            output_dir_path: Optional directory for the final processed Excel file.
                If omitted, DEFAULT_OUTPUT_DIR is used.
            poll_timeout_seconds: Timeout used for each Kafka poll.
            **kwargs: Additional arguments passed to BaseHandler.
        """
        super().__init__(**kwargs)

        self.bootstrap_servers = bootstrap_servers or BOOTSTRAP_SERVERS
        self.topic_name = topic_name or TOPIC_NAME
        self.consumer_group_id = consumer_group_id or CONSUMER_GROUP_ID
        self.output_dir_path = output_dir_path
        self.poll_timeout_seconds = poll_timeout_seconds

        if not self.bootstrap_servers:
            raise ValueError("Kafka bootstrap servers are required")
        if not self.topic_name:
            raise ValueError("Kafka topic name is required")
        if not self.consumer_group_id:
            raise ValueError("Kafka consumer group ID is required")

        self.evaluator = PolicyEvaluator(llm=llm, policy_path=policy_path)
        self.workbook = WorkbookService()
        self.consumer: Optional[Consumer] = None

        self._reset_upload_state()

    def _reset_upload_state(self) -> None:
        """Reset upload-level state without changing handler configuration."""
        self.current_upload_id: Optional[str] = None
        self.source_file_path: Optional[str] = None
        self.output_file_path: Optional[str] = None
        self.total_batches: int = 0
        self.total_records: int = 0
        self.processed_batch_ids: set[int] = set()
        self.processed_records: int = 0
        self.all_results: list[dict] = []
        self.all_work_done: bool = False

    def _build_output_file_path(self) -> str:
        """Build output path from output_dir_path or DEFAULT_OUTPUT_DIR."""
        if not self.source_file_path:
            raise ValueError("source_file_path is not available")

        source_path = Path(self.source_file_path)
        output_dir = Path(self.output_dir_path or DEFAULT_OUTPUT_DIR)
        output_dir.mkdir(parents=True, exist_ok=True)
        return str(output_dir / f"{source_path.stem}_processed.xlsx")

    def _create_consumer(self) -> Consumer:
        """
        Create and subscribe the Confluent Kafka consumer.

        Automatic offset commit is disabled so a message is committed only
        after the complete batch has been processed.
        """
        consumer = Consumer(
            {
                "bootstrap.servers": self.bootstrap_servers,
                "group.id": self.consumer_group_id,
                "auto.offset.reset": "earliest",
                "enable.auto.commit": False,
                "max.poll.interval.ms": 900000,
                "session.timeout.ms": 45000,
            }
        )
        consumer.subscribe([self.topic_name])
        return consumer

    @staticmethod
    def _is_noise_column(key: Any) -> bool:
        """
        Return True for generated/internal columns that should not be sent to
        the policy evaluator.
        """
        text = str(key).strip()

        if not text:
            return True
        if text in RESULT_HEADERS:
            return True
        if text.lower().startswith("unnamed"):
            return True

        return False

    async def _process_record(
        self,
        record: dict,
        *,
        upload_id: str,
        batch_id: int,
    ) -> dict:
        """
        Evaluate one transaction and return its result without modifying Excel.

        Args:
            record: Transaction dictionary received from Kafka.
            upload_id: Upload identifier.
            batch_id: Current batch identifier.

        Returns:
            Dictionary containing the row number and evaluation result.
        """
        row_number = record.get(ROW_NUMBER_FIELD)

        clean_record = {
            key: value
            for key, value in record.items()
            if key != ROW_NUMBER_FIELD and not self._is_noise_column(key)
        }

        logger.info("--------------------------------------------")
        logger.info(
            "Evaluating upload=%s batch=%s row=%s",
            upload_id,
            batch_id,
            row_number,
        )

        try:
            result = await self.evaluator.evaluate(clean_record)

            logger.info(
                "Evaluation result row=%s decision=%s score=%s severity=%s policy=%s",
                row_number,
                result.get("decision"),
                result.get("threat_score"),
                result.get("threat_severity"),
                result.get("policy_sid"),
            )

        except Exception as exc:
            logger.exception("Evaluation failed row=%s", row_number)

            result = {
                "decision": "ERROR",
                "threat_score": None,
                "threat_severity": None,
                "threat_reasons": [str(exc)],
                "policy_sid": None,
                "matched_policy_sids": [],
            }


        self.processed_records += 1

        logger.info(
            "Record complete %s/%s",
            self.processed_records,
            self.total_records,
        )

        return {
            "row_number": row_number,
            **result,
        }

    async def _process_batch(self, batch: dict) -> list[dict]:
        """
        Process every valid transaction record in one Kafka batch.

        Args:
            batch: Kafka batch payload.

        Returns:
            List of per-record evaluation results.

        Raises:
            ValueError: If the batch records field is not a list.
        """
        upload_id = str(batch.get("upload_id") or "")
        batch_id = int(batch.get("batch_id") or 0)
        file_path = batch.get("source_file_path")
        records = batch.get("records") or []

        if not isinstance(records, list):
            raise ValueError("Kafka batch 'records' must be a list")

        if file_path:
            self.source_file_path = file_path

        logger.info("============================================")
        logger.info("Processing batch %s/%s", batch_id, self.total_batches)
        logger.info("Records in batch: %s", len(records))
        logger.info("============================================")

        results: list[dict] = []

        for record in records:
            if not isinstance(record, dict):
                logger.warning(
                    "Skipping invalid record in batch %s: %r",
                    batch_id,
                    record,
                )
                continue

            result = await self._process_record(
                record,
                upload_id=upload_id,
                batch_id=batch_id,
            )
            results.append(result)
            self.all_results.append(result)

        return results

    def _register_upload(self, batch: dict) -> None:
        """
        Register upload metadata from the first received batch.

        Args:
            batch: Kafka batch containing upload metadata.

        Raises:
            ValueError: If upload_id is missing.
        """
        upload_id = str(batch.get("upload_id") or "")

        if not upload_id:
            raise ValueError("Kafka message is missing upload_id")

        self.current_upload_id = upload_id
        self.total_batches = int(batch.get("total_batches") or 0)
        self.total_records = int(batch.get("total_data_count") or 0)
        self.source_file_path = batch.get("source_file_path")

        logger.info("============================================")
        logger.info("NEW UPLOAD")
        logger.info("Upload ID: %s", self.current_upload_id)
        logger.info("Total batches: %s", self.total_batches)
        logger.info("Total records: %s", self.total_records)
        logger.info("Excel file: %s", self.source_file_path)
        logger.info("============================================")

    def _is_upload_complete(self) -> bool:
        """
        Return True only when all expected batches and all expected records
        have been processed.
        """
        batches_done = (
            self.total_batches > 0
            and len(self.processed_batch_ids) >= self.total_batches
        )

        records_done = (
            self.total_records > 0
            and self.processed_records >= self.total_records
        )

        return batches_done and records_done

    async def _handle_message(self, message) -> dict:
        """
        Decode and process one Kafka message.

        Returns:
            Processing status for the current message.

        Raises:
            ValueError: For empty or structurally invalid messages.
            UnicodeDecodeError: For invalid UTF-8 payloads.
            json.JSONDecodeError: For invalid JSON payloads.
        """
        raw_value = message.value()

        if raw_value is None:
            raise ValueError("Kafka message value is empty")

        if isinstance(raw_value, bytes):
            raw_value = raw_value.decode("utf-8")

        batch = json.loads(raw_value)

        if not isinstance(batch, dict):
            raise ValueError("Kafka message must contain JSON object")

        upload_id = str(batch.get("upload_id") or "")
        batch_id = int(batch.get("batch_id") or 0)

        if self.current_upload_id is None:
            self._register_upload(batch)

        if upload_id != self.current_upload_id:
            logger.warning(
                "Different upload received=%s current=%s. Ignoring for this run.",
                upload_id,
                self.current_upload_id,
            )
            return {
                "status": "ignored",
                "reason": "different_upload",
                "upload_id": upload_id,
                "batch_id": batch_id,
            }

        if batch_id in self.processed_batch_ids:
            logger.warning("Duplicate batch %s ignored", batch_id)
            return {
                "status": "duplicate",
                "upload_id": upload_id,
                "batch_id": batch_id,
            }

        logger.info("Received batch %s/%s", batch_id, self.total_batches)

        results = await self._process_batch(batch)

        self.processed_batch_ids.add(batch_id)

        logger.info(
            "Completed batches: %s/%s",
            len(self.processed_batch_ids),
            self.total_batches,
        )

        return {
            "status": "processed",
            "upload_id": upload_id,
            "batch_id": batch_id,
            "records_processed": len(results),
            "results": results,
        }

    @tool
    async def consume_policy_batches(self, **kwargs) -> dict:
        """
        Consume banking policy batches from Kafka and process one full upload.

        The tool waits for Kafka messages, evaluates every transaction, collects all
        results, manually commits each successfully processed Kafka message, and
        creates a separate output workbook after the full upload is complete.

        Args:
            **kwargs: Optional SuperAgentX workflow arguments, including
                ``previous_agent_result``.

        Returns:
            Dictionary containing upload ID, file path, total/processed batch
            counts, total/processed record counts, completion status, and any
            consumer-level error.
        """
        previous_agent_result = kwargs.get("previous_agent_result")

        if previous_agent_result is not None:
            logger.info("Previous agent result received: %s", previous_agent_result)

        self._reset_upload_state()
        self.consumer = self._create_consumer()

        logger.info("Consumer started.")
        logger.info(
            "Kafka=%s Topic=%s Group=%s",
            self.bootstrap_servers,
            self.topic_name,
            self.consumer_group_id,
        )
        logger.info("Waiting for producer batches...")

        error_message: Optional[str] = None

        try:
            while True:
                message = await asyncio.to_thread(
                    self.consumer.poll,
                    self.poll_timeout_seconds,
                )

                if message is None:
                    continue

                if message.error():
                    logger.error("Kafka error: %s", message.error())
                    continue

                try:
                    result = await self._handle_message(message)

                    # Different-upload messages are intentionally not committed
                    # here so another consumer/run can process them.
                    if result.get("status") == "ignored":
                        continue

                    await asyncio.to_thread(
                        self.consumer.commit,
                        message=message,
                        asynchronous=False,
                    )

                    if self._is_upload_complete():
                        logger.info("All batches and records processed. Creating output Excel...")

                        self.output_file_path = self._build_output_file_path()

                        await asyncio.to_thread(
                            self.workbook.create_output_workbook,
                            self.source_file_path,
                            self.output_file_path,
                            self.all_results,
                        )

                        self.all_work_done = True

                        logger.info("============================================")
                        logger.info("ALL BATCHES COMPLETED — OUTPUT EXCEL CREATED")
                        logger.info("Upload ID: %s", self.current_upload_id)
                        logger.info("Source Excel: %s", self.source_file_path)
                        logger.info("Output Excel: %s", self.output_file_path)
                        logger.info("Processed batches: %s/%s", len(self.processed_batch_ids), self.total_batches)
                        logger.info("Processed records: %s/%s", self.processed_records, self.total_records)
                        logger.info("Stopping Kafka handler...")
                        logger.info("============================================")
                        break

                except Exception as exc:
                    # Failed batch is not committed.
                    logger.exception("Batch processing failed")
                    error_message = str(exc)

        except asyncio.CancelledError:
            logger.info("Kafka consumer task cancelled.")
            raise

        except KeyboardInterrupt:
            logger.info("Consumer manually stopped.")

        except Exception as exc:
            logger.exception("Kafka consumer failed")
            error_message = str(exc)

        finally:
            if self.consumer is not None:
                try:
                    await asyncio.to_thread(self.consumer.close)
                except Exception:
                    logger.exception("Error closing Kafka consumer")
                finally:
                    self.consumer = None

            logger.info("Consumer stopped.")

        return {
            "status": "completed" if self.all_work_done else "failed",
            "upload_id": self.current_upload_id,
            "source_file_path": self.source_file_path,
            "output_file_path": self.output_file_path,
            "output_dir_path": self.output_dir_path or DEFAULT_OUTPUT_DIR,
            "total_batches": self.total_batches,
            "processed_batches": len(self.processed_batch_ids),
            "total_records": self.total_records,
            "processed_records": self.processed_records,
            "output_excel_created": self.all_work_done,
            "error": error_message,
        }

    @tool
    async def get_consumer_status(self) -> dict:
        """
        Return the current Kafka consumer processing state.

        Returns:
            Dictionary containing upload, batch, record, completion, and
            running-state information.
        """
        return {
            "upload_id": self.current_upload_id,
            "source_file_path": self.source_file_path,
            "output_file_path": self.output_file_path,
            "output_dir_path": self.output_dir_path or DEFAULT_OUTPUT_DIR,
            "total_batches": self.total_batches,
            "processed_batches": len(self.processed_batch_ids),
            "total_records": self.total_records,
            "processed_records": self.processed_records,
            "completed": self.all_work_done,
            "running": self.consumer is not None,
        }


# ============================================================

# MAIN

# ============================================================

if __name__ == "__main__":

    llm_config = {
        "model": "gemini-2.5-flash",
        "llm_type": "gemini",
    }

    llm_client: LLMClient = LLMClient(
        llm_config=llm_config
    )

    handler = KafkaConsumerHandler(
        policy_path="/home/bala/Downloads/Banking-policy-evalution/Kafka-policy-evalution/policy.json",
        llm=llm_client,
        output_dir_path="/home/bala/Downloads/"
    )

    result = asyncio.run(

        handler.consume_policy_batches()

    )

    if result.get("status") == "completed":
        logger.info(

            "All data processed and Excel file fully updated. Exiting."

        )

        sys.exit(0)

    logger.warning(

        "Handler stopped before all work completed. Result=%s",

        result,

    )

    sys.exit(1)